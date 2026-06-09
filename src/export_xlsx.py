"""
Genera output/PRODE_Mundial_2026.xlsx a partir de output/picks.json.
Tres hojas: Pronosticos (para cargar), Especiales, Detalle tecnico.
Corré primero src/model.py.
"""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from data import GRUPOS

OUT = os.path.join(os.path.dirname(__file__), "..", "output")
rows = json.load(open(os.path.join(OUT, "picks.json"), encoding="utf-8"))

NAVY, BLUE, GREY, GOLD = "1F3864", "2E5496", "F2F2F2", "FFE699"
hf = Font(name="Arial", bold=True, color="FFFFFF", size=11)
nf = Font(name="Arial", size=10)
bf = Font(name="Arial", bold=True, size=10)
ctr = Alignment(horizontal="center", vertical="center")
lft = Alignment(horizontal="left", vertical="center")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = "Pronosticos"
ws.merge_cells("A1:F1")
ws["A1"] = "PRODE MUNDIAL 2026 - Fase de grupos (modelo Dixon-Coles, optimo para 5 pts exacto / 2 pts signo)"
ws["A1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
ws["A1"].fill = PatternFill("solid", start_color=NAVY); ws["A1"].alignment = ctr
ws.row_dimensions[1].height = 26
for c, h in enumerate(["Grupo", "Jornada", "Partido", "MARCADOR", "Signo", "Confianza"], 1):
    cell = ws.cell(2, c, h); cell.font = hf
    cell.fill = PatternFill("solid", start_color=BLUE); cell.alignment = ctr; cell.border = border

r = 3
for g in "ABCDEFGHIJKL":
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    cell = ws.cell(r, 1, f"GRUPO {g}  -  {', '.join(GRUPOS[g])}")
    cell.font = hf; cell.fill = PatternFill("solid", start_color=NAVY); cell.alignment = lft; cell.border = border
    r += 1
    for x in [z for z in rows if z["grupo"] == g]:
        i, j = x["pick"].split("-")
        sign = "1 (local)" if int(i) > int(j) else ("2 (visita)" if int(i) < int(j) else "X (empate)")
        vals = [x["grupo"], x["jor"], f'{x["local"]} vs {x["visit"]}', x["pick"], sign, x["conf"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v); cell.font = bf if c == 4 else nf
            cell.alignment = lft if c == 3 else ctr; cell.border = border
            if c == 4: cell.fill = PatternFill("solid", start_color=GOLD)
            elif r % 2 == 0: cell.fill = PatternFill("solid", start_color=GREY)
        r += 1
for col, w in zip("ABCDEF", [7, 8, 34, 11, 13, 11]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A3"

ws2 = wb.create_sheet("Especiales")
ws2.merge_cells("A1:C1")
ws2["A1"] = "PRONOSTICOS ESPECIALES (optimos por probabilidad)"
ws2["A1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
ws2["A1"].fill = PatternFill("solid", start_color=NAVY); ws2["A1"].alignment = ctr
for c, h in enumerate(["Pronostico", "Eleccion", "Razon"], 1):
    cell = ws2.cell(2, c, h); cell.font = hf
    cell.fill = PatternFill("solid", start_color=BLUE); cell.alignment = ctr; cell.border = border
esp = [("Campeon", "Espana", "#1 en Opta (16,1%) y DTAI (24%), co-favorita de mercado"),
       ("Subcampeon", "Francia", "Co-favorita a la final (+240); cuadro opuesto a Espana"),
       ("Mas goles", "Mbappe (Francia)", "Favorito Bota de Oro (+600); 9 + penalista + recorrido largo")]
for i, t in enumerate(esp, 3):
    for c, v in enumerate(t, 1):
        cell = ws2.cell(i, c, v); cell.font = bf if c == 2 else nf
        cell.alignment = lft; cell.border = border
        if c == 2: cell.fill = PatternFill("solid", start_color=GOLD)
for col, w in zip("ABC", [22, 20, 60]):
    ws2.column_dimensions[col].width = w

ws3 = wb.create_sheet("Detalle tecnico")
hd = ["Grupo", "Jor", "Local", "Visitante", "xG L", "xG V", "P(1)%", "P(X)%", "P(2)%", "Pick", "P(ex)%", "EV"]
for c, h in enumerate(hd, 1):
    cell = ws3.cell(1, c, h); cell.font = hf
    cell.fill = PatternFill("solid", start_color=BLUE); cell.alignment = ctr; cell.border = border
for i, x in enumerate(rows, 2):
    vals = [x["grupo"], x["jor"], x["local"], x["visit"], x["lh"], x["la"],
            round(x["ph"] * 100), round(x["pd"] * 100), round(x["pa"] * 100),
            x["pick"], round(x["pexact"] * 100, 1), x["ev"]]
    for c, v in enumerate(vals, 1):
        cell = ws3.cell(i, c, v); cell.font = nf
        cell.alignment = lft if c in (3, 4) else ctr; cell.border = border
for col, w in zip("ABCDEFGHIJKL", [6, 4, 16, 16, 7, 7, 7, 7, 7, 6, 8, 6]):
    ws3.column_dimensions[col].width = w
ws3.freeze_panes = "A2"

wb.save(os.path.join(OUT, "PRODE_Mundial_2026.xlsx"))
print("Excel generado en output/PRODE_Mundial_2026.xlsx")
